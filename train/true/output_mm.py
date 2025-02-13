import os
import subprocess
import openpyxl
import argparse
from openpyxl import Workbook
import numpy as np


wb = Workbook()
sheet = wb.active
sheet.title = "Training Results"


headers = ["animals", "cities", "companies", "elements", "facts", "inventions"]
sheet.append(headers)


script_name = "train_mm.py" 
data_path = "../../hd_data_prompt/true/llama2chat7b/"
output_path = "../../output/"
os.makedirs(output_path, exist_ok=True)
base_args = [
    "python", script_name,
    "--output_path", output_path,
    "--data_path", data_path,
    "--device", "cuda:0"
]


for train_number in range(6):

    args = base_args + ["--train_number", str(train_number)]

    process = subprocess.Popen(args, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
    stdout, stderr = process.communicate()

    results = [None] * 6
    results[train_number] = None 

    lines = stdout.splitlines()
    for line in lines:
        if "Accuracy on" in line:
            try:
                parts = line.split("Accuracy on")[1].split(":")
                dataset = parts[0].strip()
                accuracy = float(parts[1].strip())
                index = headers.index(dataset)
                results[index] = round(accuracy, 4)
            except Exception as e:
                print(f"Error parsing line: {line} - {e}")

    sheet.append(results)


row_count = sheet.max_row
col_count = sheet.max_column


for row in range(2, row_count + 1):
    values = [sheet.cell(row=row, column=col).value for col in range(1, col_count + 1)]
    values = [v for v in values if v is not None]
    if values:
        row_avg = round(np.mean(values), 4)
        sheet.cell(row=row, column=col_count + 1, value=row_avg)


col_averages = []
for col in range(1, col_count + 1):
    values = [sheet.cell(row=row, column=col).value for row in range(2, row_count + 1)]
    values = [v for v in values if v is not None]
    if values:
        col_avg = round(np.mean(values), 4)
        sheet.cell(row=row_count + 1, column=col, value=col_avg)


all_values = []
for row in range(2, row_count + 1):
    for col in range(1, col_count + 1):
        value = sheet.cell(row=row, column=col).value
        if value is not None:
            all_values.append(value)


if all_values:
    overall_avg = round(np.mean(all_values),4)
    sheet.cell(row=row_count + 2, column=1, value="Overall Avg")
    sheet.cell(row=row_count + 2, column=2, value=overall_avg)


save_path = f"{output_path}/train_mm_true.xlsx"
wb.save(save_path)
print(f"Saved workbook: {save_path}")
